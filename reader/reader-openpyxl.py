import openpyxl 
from openpyxl.chart import Reference
from openpyxl.chart import BarChart

wb = openpyxl.load_workbook('assets/cars-2023.xlsx')

ws = wb.active

print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

print('The value in cell A1 is: '+ws['A1'].value)

values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)

data=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)


my_list = list()

for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)

print(my_list)    
    
# for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
#     (print ("{:<8}{:<35}{:<10}
#              {:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6)))

values = Reference(ws,         # worksheet object   
                   min_col=2,  # minimum column where your values begin
                   max_col=2,  # maximum column where your values end
                   min_row=1,  # minimum row you’d like to plot from
                   max_row=13) # maximum row you’d like to plot from

chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(values)

# set the title of the chart
chart.title = "Total Sales"

# set the title of the x-axis
chart.x_axis.title = "Genre"

# set the title of the y-axis
chart.y_axis.title = "Total Sales by Genre"

# the top-left corner of the chart
# is anchored to cell F2 .
ws.add_chart(chart,"J3")

# save the file 
wb.save("assets/cars-2023.xlsx")